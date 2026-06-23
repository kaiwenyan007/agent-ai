# -*- coding: utf-8 -*-
"""生成腾讯云 TRTC + 云直播采购清单 Excel（含码率测算）"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

OUTPUT = Path(__file__).resolve().parent.parent / "docs" / "腾讯云直播采购清单_TRTC与云直播.xlsx"

# ========== 业务参数（可修改后重新生成）==========
SESSIONS_PER_MONTH = 70
DURATION_MIN = 90
VIEWERS_LIST = [2000, 3500, 5000]
BITRATES_KBPS = [800, 1000, 1500, 2000, 2500]  # 混流输出码率档位
DEFAULT_BITRATE = 1500
LEB_PACKAGE_TB = 146  # 100TB特惠包可抵快直播 TB
LEB_PACKAGE_PRICE = 25939
TRTC_BASIC_MONTHLY = 625

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
SUB_FONT = Font(bold=True, size=11)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
REC_FILL = PatternFill("solid", fgColor="E2EFDA")
HIGHLIGHT_FILL = PatternFill("solid", fgColor="DDEBF7")


def gb_per_viewer_hour(bitrate_kbps: int) -> float:
    """单人单小时快直播流量(GB)"""
    return bitrate_kbps / 8 / 1024 * 3600 / 1024


def session_flux_gb(bitrate_kbps: int, viewers: int, duration_min: int = DURATION_MIN) -> float:
    return gb_per_viewer_hour(bitrate_kbps) * viewers * (duration_min / 60)


def monthly_flux_tb(bitrate_kbps: int, viewers: int) -> float:
    return session_flux_gb(bitrate_kbps, viewers) * SESSIONS_PER_MONTH / 1024


def estimate_postpaid_monthly_yuan(flux_tb: float) -> str:
    """粗算后付费月费区间（混合阶梯）"""
    flux_gb = flux_tb * 1024
    low = int(flux_gb * 0.42)
    high = int(flux_gb * 0.52)
    return f"{low:,}-{high:,}"


def packages_needed(flux_tb: float) -> int:
    import math
    return max(1, math.ceil(flux_tb / LEB_PACKAGE_TB))


def fmt_tb(gb: float) -> str:
    tb = gb / 1024
    if tb >= 1:
        return f"~{tb:.2f} TB"
    return f"~{gb:.0f} GB"


def style_header_row(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def write_table(ws, start_row, headers, rows, col_widths=None, highlight_rows=None):
    for i, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=i, value=h)
    style_header_row(ws, start_row, len(headers))
    highlight_rows = highlight_rows or set()
    for ri, row in enumerate(rows, start_row + 1):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if ri in highlight_rows:
                cell.fill = HIGHLIGHT_FILL
    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    return start_row + len(rows) + 2


def add_title(ws, row, text, merge_cols=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=merge_cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = TITLE_FONT
    return row + 2


def build_procurement_note(viewers: int, bitrate: int) -> tuple[int, float, int]:
    tb = monthly_flux_tb(bitrate, viewers)
    pkgs = packages_needed(tb)
    annual = pkgs * LEB_PACKAGE_PRICE + TRTC_BASIC_MONTHLY * 12
    return pkgs, tb, annual


def main():
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    default_tb = monthly_flux_tb(DEFAULT_BITRATE, 2000)
    max_tb = monthly_flux_tb(DEFAULT_BITRATE, 5000)
    pkgs_a, _, annual_a = build_procurement_note(2000, DEFAULT_BITRATE)
    pkgs_b, _, annual_b = build_procurement_note(5000, DEFAULT_BITRATE)

    # ========== Sheet1: 总览 ==========
    ws = wb.active
    ws.title = "总览"
    r = 1
    r = add_title(ws, r, "腾讯云 TRTC + 云直播 采购清单总览", 5)
    ws.cell(row=r, column=1, value="业务场景").font = SUB_FONT
    r += 1
    overview = [
        ("月直播场次", f"{SESSIONS_PER_MONTH} 场"),
        ("单场时长（默认）", f"{DURATION_MIN} 分钟"),
        ("观众规模", "2,000 ~ 5,000 人/场"),
        ("混流输出码率（默认）", f"{DEFAULT_BITRATE} kbps（720p 推荐）"),
        ("上麦人数上限", "6 人"),
        ("架构", "TRTC 互动（教师+上麦）+ 混流旁路 + 快直播 CDN 大班观看"),
        ("必开产品", "TRTC + 云直播（两个都要，不能只买云直播）"),
        ("数据说明", "刊例价参考腾讯云公开文档，流量按视频码率×人数×时长计算"),
        ("生成日期", "2026-06-18"),
    ]
    for k, v in overview:
        ws.cell(row=r, column=1, value=k).border = BORDER
        ws.cell(row=r, column=2, value=v).border = BORDER
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="推荐采购结论（含码率）").font = SUB_FONT
    r += 1
    conclusions = [
        f"1. TRTC：基础版 625元/月（旁路+混流，赠11万分钟/月）",
        f"2. 云直播：快直播特惠 100TB 包（可抵~{LEB_PACKAGE_TB}TB），折合~0.18元/GB",
        f"3. 默认 {DEFAULT_BITRATE}kbps + 2000人：月流量 ~{default_tb:.0f}TB → 建议 {pkgs_a}×100TB特惠，首年约 {annual_a:,}元",
        f"4. 默认 {DEFAULT_BITRATE}kbps + 5000人：月流量 ~{max_tb:.0f}TB → 建议 {pkgs_b}×100TB特惠，首年约 {annual_b:,}元",
        "5. 码率每提高 33%，CDN 流量同比例增加；1000kbps 比 1500kbps 省约 33%",
        "6. 云直播计费方式必须为「日结流量」，详见「码率与流量测算」Sheet",
    ]
    for line in conclusions:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        c = ws.cell(row=r, column=1, value=line)
        c.fill = REC_FILL
        c.alignment = Alignment(wrap_text=True)
        r += 1
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 72

    # ========== Sheet2: 码率与流量测算（新增/核心）==========
    ws_bit = wb.create_sheet("码率与流量测算", 1)
    r = add_title(ws_bit, 1, "视频码率 → 快直播 CDN 流量测算", 8)

    ws_bit.cell(row=r, column=1, value="流量计算公式").font = SUB_FONT
    r += 1
    formulas = [
        "单场流量(GB) = 视频码率(kbps) ÷ 8 ÷ 1024 × 时长(秒) × 观众人数 ÷ 1024",
        f"简化为：单场流量(GB) = 码率(kbps) × 0.000675 × 时长(小时) × 观众人数",
        f"示例：1500kbps × 0.000675 × 1.5h × 2000人 = {session_flux_gb(1500, 2000):.0f} GB/场",
        "说明：码率为混流输出总码率（含音视频），CDN 计费按下行流量，与协议(webrtc/flv)无关",
    ]
    for f in formulas:
        ws_bit.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        ws_bit.cell(row=r, column=1, value=f).fill = NOTE_FILL
        r += 1
    r += 1

    ws_bit.cell(row=r, column=1, value="推荐混流码率参考（education）").font = SUB_FONT
    r += 1
    r = write_table(
        ws_bit, r,
        ["输出分辨率", "推荐码率(kbps)", "单场2000人90min", "月70场流量", "画质", "适用场景"],
        [
            ["540p", 800, fmt_tb(session_flux_gb(800, 2000)), f"~{monthly_flux_tb(800, 2000):.0f} TB", "一般", "省流量/弱网"],
            ["720p", 1000, fmt_tb(session_flux_gb(1000, 2000)), f"~{monthly_flux_tb(1000, 2000):.0f} TB", "良好", "大班课推荐(省)"],
            ["720p", 1500, fmt_tb(session_flux_gb(1500, 2000)), f"~{monthly_flux_tb(1500, 2000):.0f} TB", "较好", "默认估算档位"],
            ["720p", 2000, fmt_tb(session_flux_gb(2000, 2000)), f"~{monthly_flux_tb(2000, 2000):.0f} TB", "好", "课件细节多"],
            ["1080p", 2500, fmt_tb(session_flux_gb(2500, 2000)), f"~{monthly_flux_tb(2500, 2000):.0f} TB", "很好", "高清要求/慎用"],
        ],
        [12, 14, 16, 14, 10, 22],
        highlight_rows={r + 3},  # highlight 1500kbps row - need fix
    )

    ws_bit.cell(row=r, column=1, value="码率 × 观众 月流量矩阵（TB/月，70场×90min）").font = SUB_FONT
    r += 1
    matrix_headers = ["码率(kbps)"] + [f"{v}人/场" for v in VIEWERS_LIST] + ["100TB包数(约)", "后付费估算(元/月,2000人)"]
    matrix_rows = []
    highlight_matrix_row = None
    for br in BITRATES_KBPS:
        row = [br]
        for v in VIEWERS_LIST:
            row.append(round(monthly_flux_tb(br, v), 1))
        row.append(packages_needed(monthly_flux_tb(br, 2000)))
        row.append(estimate_postpaid_monthly_yuan(monthly_flux_tb(br, 2000)))
        matrix_rows.append(row)
        if br == DEFAULT_BITRATE:
            highlight_matrix_row = r + len(matrix_rows)
    r = write_table(ws_bit, r, matrix_headers, matrix_rows, [12, 12, 12, 12, 14, 22],
                    highlight_rows={highlight_matrix_row} if highlight_matrix_row else None)

    ws_bit.cell(row=r, column=1, value="单场流量明细（GB/场，90分钟）").font = SUB_FONT
    r += 1
    detail_headers = ["码率(kbps)"] + [f"{v}人" for v in VIEWERS_LIST]
    detail_rows = [[br] + [round(session_flux_gb(br, v), 0) for v in VIEWERS_LIST] for br in BITRATES_KBPS]
    hl = r + BITRATES_KBPS.index(DEFAULT_BITRATE) + 1
    r = write_table(ws_bit, r, detail_headers, detail_rows, [12] + [12] * len(VIEWERS_LIST), highlight_rows={hl})

    ws_bit.cell(row=r, column=1, value="码率调整系数（相对1500kbps）").font = SUB_FONT
    r += 1
    r = write_table(
        ws_bit, r,
        ["码率(kbps)", "相对1500kbps系数", "2000人月流量(TB)", "采购影响"],
        [
            [800, round(800 / 1500, 2), round(monthly_flux_tb(800, 2000), 1), "流量最低，画质一般"],
            [1000, round(1000 / 1500, 2), round(monthly_flux_tb(1000, 2000), 1), "推荐省流方案"],
            [1500, 1.0, round(monthly_flux_tb(1500, 2000), 1), "默认基准"],
            [2000, round(2000 / 1500, 2), round(monthly_flux_tb(2000, 2000), 1), "流量+33%"],
            [2500, round(2500 / 1500, 2), round(monthly_flux_tb(2500, 2000), 1), "流量+67%，高清慎用"],
        ],
        [12, 18, 18, 28],
        highlight_rows={r + 3},
    )

    # ========== Sheet3: TRTC套餐 ==========
    ws2 = wb.create_sheet("TRTC收费标准")
    r = add_title(ws2, 1, "TRTC 包月套餐（按 SdkAppId）", 7)
    r = write_table(
        ws2, r,
        ["版本", "月费(元)", "赠送音视频时长/月", "旁路推流", "云端混流", "适合场景", "备注"],
        [
            ["入门版", 0, "无", "否", "否", "不适用", "无法旁路混流"],
            ["基础版", 625, "11万分钟", "是", "是", "推荐（最低配）", "解锁旁路+混流"],
            ["尊享版", 1875, "38万分钟", "是", "是", "互动用量大", ""],
            ["尊享版 Plus", 2875, "38万分钟+26万转码分钟", "是", "是", "混流/录制多", "限时8折2875"],
            ["旗舰版", 6250, "140万分钟", "是", "是", "一般不必", ""],
            ["旗舰版 Plus", 8000, "140万分钟+60万转码分钟", "是", "是", "一般不必", "限时8折8000"],
        ],
        [12, 12, 22, 10, 10, 18, 18],
    )
    ws2.cell(row=r, column=1, value="TRTC 后付费 - 音视频时长").font = SUB_FONT
    r += 1
    r = write_table(
        ws2, r,
        ["档位", "分辨率", "单价(元/千分钟)"],
        [
            ["音频", "纯音频", 7.0],
            ["标清 SD", "≤640×480", 14.0],
            ["高清 HD", "640×480~1280×720", 28.0],
            ["超高清 FHD", "1280×720~1920×1080", 63.0],
            ["2K", "1920×1080~2560×1440", 112.0],
            ["4K", "2560×1440~4096×2176", 252.0],
        ],
        [14, 28, 18],
    )
    ws2.cell(row=r, column=1, value="TRTC 后付费 - 增值能力").font = SUB_FONT
    r += 1
    r = write_table(
        ws2, r,
        ["计费项", "单价", "说明"],
        [
            ["云端混流-音频输入", "5.6 元/千分钟", "按混流输入时长"],
            ["云端混流-HD输入(H.264)", "21 元/千分钟", "720p混流输入"],
            ["云端混流-SD输入(H.264)", "12 元/千分钟", ""],
            ["云端混流-FHD输入(H.264)", "48 元/千分钟", ""],
            ["旁路转推", "8 元/千分钟", "2025-10-23后新账号按时长计费"],
        ],
        [22, 18, 35],
    )
    ws2.cell(row=r, column=1, value="TRTC 音视频通用套餐包（连麦包同款）").font = SUB_FONT
    r += 1
    r = write_table(
        ws2, r,
        ["规格(千分钟)", "价格(元)", "折合单价(元/千分钟)", "抵扣比例说明"],
        [
            [25, 168, 6.72, "语音:标清:高清:全高清 = 1:2:4:15"],
            [250, 1588, 6.35, "1分钟HD扣4分钟包时长"],
            [1000, 5968, 5.97, ""],
            [3000, 16888, 5.63, ""],
        ],
        [16, 12, 22, 35],
    )

    # ========== Sheet4: 云直播套餐 ==========
    ws3 = wb.create_sheet("云直播收费标准")
    r = add_title(ws3, 1, "云直播 快直播(LEB) 后付费 - 日结流量", 4)
    r = write_table(
        ws3, r,
        ["日流量阶梯", "单价(元/GB)"],
        [
            ["0 - 2 TB", 0.52],
            ["2 - 10 TB", 0.50],
            ["10 - 50 TB", 0.48],
            ["50 - 100 TB", 0.44],
            ["100 TB - 1 PB", 0.38],
            ["≥ 1 PB", 0.32],
        ],
        [20, 15],
    )
    ws3.cell(row=r, column=1, value="云直播 直播流量资源包（普通）").font = SUB_FONT
    r += 1
    r = write_table(
        ws3, r,
        ["规格", "价格(元)", "可抵快直播(约)", "备注"],
        [
            ["100 GB", 26, "~50 GB", "快直播抵扣比 1:2"],
            ["500 GB", 128, "~250 GB", ""],
            ["1 TB", 248, "~500 GB", ""],
            ["5 TB", 1200, "~2.5 TB", ""],
            ["10 TB", 2350, "~5 TB", ""],
            ["50 TB", 9889, "~25 TB", ""],
            ["200 TB", 35500, "~100 TB", ""],
        ],
        [12, 12, 16, 25],
    )
    ws3.cell(row=r, column=1, value="云直播 快直播特惠流量包（★推荐）").font = SUB_FONT
    r += 1
    r = write_table(
        ws3, r,
        ["包规格", "价格(元)", "可抵快直播(中国境内)", "折合快直播单价", "备注"],
        [
            ["73 GB", 19, "~50 GB", "~0.38 元/GB", "抵扣比 1:1.46"],
            ["2.5 TB", 878, "~3.65 TB", "~0.24 元/GB", ""],
            ["5 TB", 1720, "~7.3 TB", "~0.24 元/GB", ""],
            ["7 TB", 2310, "~10.22 TB", "~0.23 元/GB", ""],
            ["100 TB", 25939, f"~{LEB_PACKAGE_TB} TB", "~0.18 元/GB", "主力采购档"],
        ],
        [12, 12, 22, 16, 25],
    )
    ws3.cell(row=r, column=1, value="重要说明").font = SUB_FONT
    r += 1
    notes = [
        "CDN流量与视频码率成正比：码率越高，同等人数下流量包消耗越快",
        "资源包有效期一般为1年，仅支持「日结流量」计费方式",
        "快直播特惠包比普通流量包更划算（快直播抵扣 1:1.46 vs 普通 1:2）",
    ]
    for n in notes:
        ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        ws3.cell(row=r, column=1, value=n).fill = NOTE_FILL
        r += 1

    # ========== Sheet5: 用量测算 ==========
    ws4 = wb.create_sheet("用量测算")
    r = add_title(ws4, 1, "业务量假设与用量测算", 6)
    r = write_table(
        ws4, r,
        ["参数", "取值", "说明"],
        [
            ["月直播场次", SESSIONS_PER_MONTH, ""],
            ["单场时长(分钟)", DURATION_MIN, "可按比例调整"],
            ["观众规模", "2000 / 3500 / 5000", ""],
            ["混流输出码率(默认)", f"{DEFAULT_BITRATE} kbps", "见码率Sheet"],
            ["上麦人数上限", 6, ""],
            ["架构", "观众CDN观看，仅教师+上麦在TRTC", ""],
        ],
        [22, 18, 35],
    )
    ws4.cell(row=r, column=1, value=f"CDN流量测算（默认 {DEFAULT_BITRATE}kbps）").font = SUB_FONT
    r += 1
    cdn_rows = []
    hl_cdn = None
    for i, v in enumerate(VIEWERS_LIST):
        gb = session_flux_gb(DEFAULT_BITRATE, v)
        tb = monthly_flux_tb(DEFAULT_BITRATE, v)
        cdn_rows.append([v, fmt_tb(gb), f"~{tb:.0f} TB", estimate_postpaid_monthly_yuan(tb), f"需{packages_needed(tb)}×100TB包"])
    r = write_table(
        ws4, r,
        ["观众/场", "单场流量", "月70场总流量", "后付费估算(元/月)", "流量包建议"],
        cdn_rows,
        [12, 14, 16, 20, 18],
    )
    ws4.cell(row=r, column=1, value="TRTC 互动端用量（与码率无关，观众不进TRTC房）").font = SUB_FONT
    r += 1
    r = write_table(
        ws4, r,
        ["角色", "用量/场", "月70场合计(通用包分钟)", "说明"],
        [
            ["教师 HD 90min", 360, 25200, "HD按1:4扣包"],
            ["上麦3人×30min(2视频+1音频)", 240, 16800, "平均值"],
            ["混流机器人等", 100, 7000, "估算"],
            ["合计", "~700/场", "~49000/月", "远低于基础版11万分钟赠送"],
        ],
        [22, 14, 22, 25],
    )
    ws4.cell(row=r, column=1, value="混流+旁路后付费（月70场，与码率弱相关）").font = SUB_FONT
    r += 1
    r = write_table(
        ws4, r,
        ["项目", "月估算(元)"],
        [["云端混流转码", "200-400"], ["旁路转推", "50-100"]],
        [22, 18],
    )
    r += 1
    ws4.cell(row=r, column=1, value="时长调整系数（相对90分钟，码率不变）").font = SUB_FONT
    r += 1
    r = write_table(
        ws4, r,
        ["单场时长", "系数", f"2000人/{DEFAULT_BITRATE}kbps月流量"],
        [
            [60, 0.67, f"~{monthly_flux_tb(DEFAULT_BITRATE, 2000) * 0.67:.0f} TB"],
            [90, 1.0, f"~{monthly_flux_tb(DEFAULT_BITRATE, 2000):.0f} TB"],
            [120, 1.33, f"~{monthly_flux_tb(DEFAULT_BITRATE, 2000) * 1.33:.0f} TB"],
        ],
        [14, 10, 22],
    )

    # ========== Sheet6: 采购清单（含码率场景）==========
    ws5 = wb.create_sheet("采购清单")

    def proc_row(label, viewers, bitrate):
        tb = monthly_flux_tb(bitrate, viewers)
        pkgs = packages_needed(tb)
        annual = pkgs * LEB_PACKAGE_PRICE + TRTC_BASIC_MONTHLY * 12
        monthly_pkg = annual / 12
        return [
            label,
            f"{viewers}人/{bitrate}kbps",
            f"~{tb:.0f}TB/月",
            f"{pkgs}×100TB特惠",
            f"{annual:,}元/年",
            f"~{monthly_pkg:,.0f}元/月",
            f"码率{bitrate} 单场{fmt_tb(session_flux_gb(bitrate, viewers))}",
        ]

    r = add_title(ws5, 1, "按码率+人数自动测算的采购建议", 7)
    r = write_table(
        ws5, r,
        ["场景", "参数", "月流量", "建议采购", "首年预算", "月均", "备注"],
        [
            proc_row("方案A-省流", 2000, 1000),
            proc_row("方案A-默认", 2000, 1500),
            proc_row("方案A-高清", 2000, 2000),
            proc_row("方案B-默认", 3500, 1500),
            proc_row("方案B-峰值", 5000, 1500),
            proc_row("方案B-峰值高清", 5000, 2000),
        ],
        [14, 18, 12, 16, 14, 12, 22],
        highlight_rows={r + 2, r + 5},
    )

    r = add_title(ws5, r, "固定采购清单 - 方案A：2000人/1500kbps", 8)
    pkgs, tb, annual = build_procurement_note(2000, 1500)
    r = write_table(
        ws5, r,
        ["序号", "产品", "规格", "数量", "刊例价", "用途", "优先级", "备注"],
        [
            [1, "TRTC包月套餐", "基础版", "1×SdkAppId/月", "625元/月", "旁路+混流+赠11万分钟", "必买", ""],
            [2, "快直播特惠流量包", "100 TB", f"{pkgs}个/年", f"{pkgs * LEB_PACKAGE_PRICE:,}元/年", f"抵~{pkgs*LEB_PACKAGE_TB}TB", "必买", f"覆盖~{tb:.0f}TB/月"],
            [3, "快直播特惠流量包", "2.5 TB", "1个/年(缓冲)", "878元/年", "峰值/加时", "建议", ""],
            ["", "首年合计", "", "", f"约{annual + 878:,}元", "含TRTC625×12", "", f"月均~{(annual+878)/12:,.0f}元"],
        ],
        [6, 18, 14, 14, 16, 28, 10, 18],
    )

    pkgs_b, tb_b, annual_b = build_procurement_note(5000, 1500)
    ws5.cell(row=r, column=1, value="固定采购清单 - 方案B：5000人/1500kbps（★推荐）").font = SUB_FONT
    r += 1
    r = write_table(
        ws5, r,
        ["序号", "产品", "规格", "数量", "刊例价", "用途", "优先级", "备注"],
        [
            [1, "TRTC包月套餐", "基础版", "1/月", "625元/月", "旁路+混流", "必买", ""],
            [2, "快直播特惠流量包", "100 TB", f"{pkgs_b}个/年", f"{pkgs_b * LEB_PACKAGE_PRICE:,}元/年", f"抵~{pkgs_b*LEB_PACKAGE_TB}TB", "必买", f"覆盖~{tb_b:.0f}TB/月"],
            [3, "快直播特惠流量包", "7 TB档", "1个/年", "2310元/年", "补峰值", "建议", ""],
            ["", "首年合计", "", "", f"约{annual_b + 2310:,}元", "含TRTC7500", "", f"月均~{(annual_b+2310)/12:,.0f}元"],
        ],
        [6, 18, 14, 14, 16, 28, 10, 18],
    )

    ws5.cell(row=r, column=1, value="不必购买").font = SUB_FONT
    r += 1
    r = write_table(
        ws5, r,
        ["项目", "原因"],
        [
            ["仅云直播（不开TRTC）", "无法混流/连麦/旁路"],
            ["TRTC通用时长大包", "基础版赠送已够用"],
            ["普通直播流量包(非特惠)", "快直播特惠更划算"],
            ["盲目提高码率到2500kbps+", "流量增加67%+，大班课建议1000-1500kbps"],
        ],
        [28, 50],
    )

    # ========== Sheet7: 运维Checklist ==========
    ws6 = wb.create_sheet("运维Checklist")
    r = add_title(ws6, 1, "运维配置 Checklist", 4)
    r = write_table(
        ws6, r,
        ["序号", "事项", "状态(待办/完成)", "备注"],
        [
            [1, "开通 TRTC + 云直播两个产品", "待办", ""],
            [2, "TRTC应用购买「基础版」包月", "待办", "解锁旁路+混流"],
            [3, "TRTC控制台开启旁路转推+指定流旁路", "待办", ""],
            [4, "云直播配置推流域名+播放域名(备案+CNAME)", "待办", ""],
            [5, "云直播计费方式改为「日结流量」", "待办", "否则流量包不可用"],
            [6, "购买快直播特惠流量包", "待办", "按码率Sheet测算规格"],
            [7, "混流encoding.videoBitrate设为1000-1500kbps", "待办", "控制CDN流量"],
            [8, "后端streamId/roomId带事业部前缀", "待办", "便于分账"],
            [9, "设置流量包余量告警(<20%续买)", "待办", ""],
            [10, "学员端TCPlayer+webrtc://，iOS点击播放", "待办", ""],
        ],
        [6, 45, 14, 30],
    )

    # ========== Sheet8: 费用结构 ==========
    ws7 = wb.create_sheet("费用结构对比")
    r = add_title(ws7, 1, "纯TRTC vs TRTC+快直播（2000人,90min,1500kbps）", 5)
    r = write_table(
        ws7, r,
        ["方案", "单场费用(元)", "月70场(元)", "可行性", "延时"],
        [
            ["纯TRTC(2000人全进房)", "~3360", "~235200", "不可行", "300ms-1s"],
            ["TRTC+快直播CDN(推荐)", "~700", "~49000", "推荐", "500ms-1s"],
            ["节省比例", "~79%", "~79%", "", ""],
        ],
        [28, 16, 16, 28, 14],
    )
    r += 1
    ws7.cell(row=r, column=1, value="码率对CDN费用影响（2000人/场，70场/月）").font = SUB_FONT
    r += 1
    br_rows = []
    base_cost = monthly_flux_tb(1500, 2000) * 0.18 * 1000  # rough package unit cost per TB
    for br in BITRATES_KBPS:
        tb = monthly_flux_tb(br, 2000)
        pkg_cost = packages_needed(tb) * LEB_PACKAGE_PRICE / 12
        br_rows.append([f"{br} kbps", f"~{tb:.0f} TB", f"{pkgs if (pkgs := packages_needed(tb)) else 1}×100TB包", f"~{pkg_cost:,.0f}元/月(包)"])
    r = write_table(
        ws7, r,
        ["混流码率", "月CDN流量", "流量包", "包月成本粗算"],
        br_rows,
        [12, 14, 14, 18],
        highlight_rows={r + BITRATES_KBPS.index(1500) + 1},
    )

    wb.save(OUTPUT)
    print(f"Generated: {OUTPUT}")


if __name__ == "__main__":
    main()
